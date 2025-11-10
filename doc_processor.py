import os
import pypdf
import docx
import openpyxl
from bs4 import BeautifulSoup

def read_text_from_file(file_path, mime_type=None):
    """Reads content from various document types (PDF, DOCX, XLSX, TXT, HTML)."""
    # Usamos la extensión para determinar el tipo, ya que la descarga de Drive ya la definió.
    extension = os.path.splitext(file_path)[1].lower()

    try:
        if extension == '.pdf':
            with open(file_path, 'rb') as file:
                reader = pypdf.PdfReader(file)
                text = "".join(page.extract_text() or "" for page in reader.pages)
            return text
        
        elif extension == '.docx':
            doc = docx.Document(file_path)
            return "\n".join([paragraph.text for paragraph in doc.paragraphs])
            
        elif extension == '.xlsx':
            workbook = openpyxl.load_workbook(file_path)
            text = []
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text.append(f"--- Hoja: {sheet_name} ---")
                for row in sheet.iter_rows():
                    row_values = [str(cell.value) if cell.value is not None else "" for cell in row]
                    text.append("\t".join(row_values))
            return "\n".join(text)

        elif extension in ['.txt', '.html'] or mime_type in ['text/plain', 'text/html']:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
                if extension == '.html':
                    soup = BeautifulSoup(content, 'html.parser')
                    return soup.get_text(separator='\n')
                return content
        else:
            print(f"Advertencia: Tipo de archivo no soportado para lectura: {file_path}")
            return None
    except Exception as e:
        print(f"Error leyendo {file_path}: {e}")
        return None

def chunk_text(text, chunk_size=1000, chunk_overlap=100):
    """Divide el texto en fragmentos (chunks) con superposición para embeddings."""
    if not text:
        return []
    
    # Dividimos por espacios o saltos de línea y usamos tokens (palabras) para chunking
    words = text.split() 
    chunks = []
    
    # Aseguramos que la superposición no sea mayor que el chunk_size
    step = chunk_size - chunk_overlap
    if step <= 0:
        step = 1
        
    for i in range(0, len(words), step):
        chunk = " ".join(words[i:i + chunk_size])
        chunks.append(chunk)
    return chunks